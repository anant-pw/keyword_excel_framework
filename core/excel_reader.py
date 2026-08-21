"""
Reads the "TestSteps" sheet of the keyword-driven Excel workbook into
structured TestCase objects, each holding an ordered list of TestStep rows.

Column schema - matches TC.xlsx's actual semantics:
    TestCase ID | Test Scenario | Test Case Description | Keyword/Action |
    LocatorType | LocatorValue | Test Data | Expected Output | Suite

- TestCase ID: a flat, sequential integer for the ROW (like TC.xlsx) - not
  a group key. Sort order = execution order.
- Test Scenario: the grouping key. All rows sharing the same Test Scenario
  string form one logical test case (browser context, report card,
  fail-fast/continue unit).
- Test Case Description: per-ROW narrative of what that step does (e.g.
  "Enter QA email") - shown in the report so a reader can follow the flow
  without decoding Keyword/LocatorValue/TestData.
- Suite: per-STEP, not per-case. Smoke | Sanity | Regression, comma-separated
  if a step belongs to more than one (e.g. "Smoke,Sanity,Regression" for a
  shared setup step). A step only runs when the run's requested suite is in
  its Suite list - this replaces a separate Enabled column: leaving Suite
  blank, or leaving the current suite out of the list, is how you "disable"
  a step for a given run. Blank Suite means "never runs" - it does not mean
  "runs for every suite"; use an explicit comma-list for that.
- LocatorType / LocatorValue: see core/locator_resolver.py, including the
  "any" JSON-fallback-chain strategy.
- LocatorValue / Test Data may reference $name, resolved at run time from
  whatever properties/*.properties file the case's LoadProperties step
  loaded, OR from a value an earlier step in the same case captured via
  SaveAs (see core/properties_loader.py).
- SaveAs: OPTIONAL column, absent from older sheets without breaking them
  (defaults to "" per row when the column doesn't exist at all). When set
  on a row, that step's result is captured under this name for $name
  resolution by later steps in the SAME case (including inside composite
  keywords, which share the calling case's CasePropertyStore). Only a
  fixed set of keywords know how to produce a capturable result - see
  SAVE_AS_CAPABLE below; using SaveAs on any other keyword is a sheet
  error caught here at load time, not a silent no-op at run time.
- DataSource: OPTIONAL column, absent from older sheets without breaking
  them (same pattern as SaveAs). Only the FIRST row of a Test Scenario is
  consulted - it's a scenario-level attribute, not a per-step one, so
  later rows don't need it repeated (unlike Suite). When set to a sheet
  name (e.g. "LoginCreds"), that scenario is data-driven: this function
  loads the named sheet as a data table (see _load_data_table below) and
  expands the ONE scenario definition into N TestCase objects, one per
  data row - identical steps, different captured $variables. Each
  expanded case's test_scenario becomes "<original> [<Label>]", so the
  report shows which data row a failure belongs to instead of a bare
  index. A data table sheet's own header row becomes the $variable names
  available to that case's steps (via CasePropertyStore.capture, the
  same mechanism SaveAs already uses) - no new $-syntax, no new resolver.
  A data table MUST have a "Label" column; every other column becomes a
  variable. Suite filtering still applies per the base scenario's step
  Suite tags - unaffected by which/how-many data rows exist, since every
  expanded case shares the exact same step list (and therefore Suite
  values); only the captured variables differ per case.
"""
from dataclasses import dataclass, field
from pathlib import Path
from collections import OrderedDict
import re
import openpyxl

from core.exceptions import InvalidTestDataError
from core.logger import get_logger

logger = get_logger("excel_reader")

REQUIRED_COLUMNS = [
    "TestCase ID", "Test Scenario", "Test Case Description",
    "Keyword/Action", "LocatorType", "LocatorValue", "Test Data",
    "Expected Output", "Suite",
]

# Explicit opt-IN list: only these keywords require a LocatorValue. Deliberately
# an opt-in list rather than an opt-out one - a composite (common/) keyword
# name is never in here, so calling one never trips a false "requires
# LocatorValue" error just because the framework doesn't recognize the name
# yet at parse time (composite registration happens separately, before the
# main sheet is read - see core/common_keywords.py).
LOCATOR_REQUIRED = {
    "click", "clickbyjavascript", "type", "typeslowly", "clear", "hover",
    "selectdropdown", "check", "uncheck", "scrollintoview",
    "elementdisplayed", "elementnotdisplayed", "elementenabled",
    "verifytext", "verifytextcontains", "waitforelement",
    "savetext", "saveattribute", "savevalue",
    "apiget", "apipost", "apiput", "apipatch", "apidelete",  # LocatorValue holds the request URL
}

# Only these keywords produce a result worth capturing. Opt-in (not
# opt-out) for the same reason LOCATOR_REQUIRED is: a SaveAs typo or a
# SaveAs left on the wrong row (e.g. on a Click) should fail loudly at
# sheet-load time, before any browser opens, rather than silently doing
# nothing and leaving a later $name reference to fail confusingly instead.
SAVE_AS_CAPABLE = {
    "savetext", "saveattribute", "savevalue", "generatevalue",
    "apiget", "apipost", "apiput", "apipatch", "apidelete",  # optional whole-body capture
    "savejsonpath",  # mandatory - see keywords/api_keywords.py
}

# UseSession is a context-creation directive (see core/session_manager.py) -
# it must be the FIRST step of its Test Scenario, since the runner has to
# read it before opening the browser context, not while stepping through
# the case. Validated here (per-case, after grouping) rather than in
# LOCATOR_REQUIRED/SAVE_AS_CAPABLE since it's a position constraint, not a
# column-presence constraint.
SESSION_DIRECTIVE_KEYWORD = "usesession"


@dataclass
class TestStep:
    row_id: int
    test_scenario: str
    description: str
    keyword: str
    locator_type: str
    locator_value: str
    test_data: str
    expected_output: str
    suite: str
    save_as: str = ""
    data_source: str = ""


@dataclass
class TestCase:
    test_scenario: str
    steps: list = field(default_factory=list)
    data_row: dict = field(default_factory=dict)  # {} for a non-data-driven case


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _suite_list(suite_cell: str) -> list:
    """Splits a comma-separated Suite cell into a normalized (lowercase,
    trimmed) list, e.g. "Smoke, Sanity" -> ["smoke", "sanity"]. Blank -> []."""
    if not suite_cell:
        return []
    return [s.strip().lower() for s in suite_cell.split(",") if s.strip()]


_DATA_VAR_NAME = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def _load_data_table(wb, sheet_name: str) -> list:
    """Loads a DataSource sheet into a list of dicts, one per data row,
    e.g. {'Label': 'Valid login', 'username': 'demo', 'password': 'demo123'}.
    Every column except Label becomes a $variable captured into the
    CasePropertyStore for that iteration (see tests/runner.py). Fails
    loudly on any structural problem - a typo'd variable name or a missing
    Label should never surface as a confusing failure three steps into a
    run; it should surface here, at load time, before any browser opens."""
    if sheet_name not in wb.sheetnames:
        raise InvalidTestDataError(
            f"DataSource '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise InvalidTestDataError(f"DataSource sheet '{sheet_name}' is empty")

    header = [str(h).strip() if h else "" for h in rows[0]]
    if "Label" not in header:
        raise InvalidTestDataError(
            f"DataSource sheet '{sheet_name}' must have a 'Label' column identifying each "
            f"data row for the report - none found. Columns present: {header}"
        )
    var_cols = [h for h in header if h and h != "Label"]
    if not var_cols:
        raise InvalidTestDataError(
            f"DataSource sheet '{sheet_name}' has a 'Label' column but no data columns - "
            f"nothing to capture as $variables."
        )
    for h in var_cols:
        if not _DATA_VAR_NAME.match(h):
            raise InvalidTestDataError(
                f"DataSource sheet '{sheet_name}': column '{h}' isn't a valid $variable name "
                f"(must match {_DATA_VAR_NAME.pattern}) - rename the column header."
            )

    data_rows = []
    for row_num, row in enumerate(rows[1:], start=2):
        if row is None or all(v is None for v in row):
            continue
        row_dict = {h: _clean(row[i]) for i, h in enumerate(header) if h}
        if not row_dict.get("Label"):
            raise InvalidTestDataError(
                f"DataSource sheet '{sheet_name}', row {row_num}: Label is required on every row."
            )
        data_rows.append(row_dict)

    if not data_rows:
        raise InvalidTestDataError(f"DataSource sheet '{sheet_name}' has a header but zero data rows")

    labels = [r["Label"] for r in data_rows]
    dupes = sorted({label for label in labels if labels.count(label) > 1})
    if dupes:
        raise InvalidTestDataError(
            f"DataSource sheet '{sheet_name}' has duplicate Label value(s): {dupes} - labels "
            f"must be unique so report entries stay distinguishable."
        )

    return data_rows


def read_test_suite(file_path: str, sheet_name: str = "TestSteps", suite_filter: str = None) -> list:
    """
    Parse the Excel workbook into a list of TestCase objects, ordered by
    first appearance of each Test Scenario. Within a case, steps keep
    sheet row order (TestCase ID ascending). If suite_filter is given
    (e.g. "Smoke"), each case's steps are first filtered down to only
    those whose Suite matches; a case with zero matching steps is dropped
    entirely from the result.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Test sheet file not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise InvalidTestDataError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise InvalidTestDataError(f"Sheet '{sheet_name}' is empty")

    header = [str(h).strip() if h else "" for h in rows[0]]
    missing = [c for c in REQUIRED_COLUMNS if c not in header]
    if missing:
        raise InvalidTestDataError(f"Missing required columns in '{sheet_name}': {missing}")
    col_idx = {name: header.index(name) for name in header if name}

    all_steps = []
    for row_num, row in enumerate(rows[1:], start=2):
        if row is None or all(v is None for v in row):
            continue

        row_id_raw = row[col_idx["TestCase ID"]]
        if row_id_raw is None:
            continue  # skip blank separator rows

        keyword = _clean(row[col_idx["Keyword/Action"]])
        locator_value = _clean(row[col_idx["LocatorValue"]])

        if keyword.lower() in LOCATOR_REQUIRED and not locator_value:
            raise InvalidTestDataError(
                f"Row {row_num}: keyword '{keyword}' requires LocatorValue but none was given "
                f"(TestCase ID={row_id_raw})"
            )

        save_as_col = col_idx.get("SaveAs")  # optional column - absent entirely on older sheets
        save_as = _clean(row[save_as_col]) if save_as_col is not None else ""
        if save_as and keyword.lower() not in SAVE_AS_CAPABLE:
            raise InvalidTestDataError(
                f"Row {row_num}: SaveAs='{save_as}' set on keyword '{keyword}', which doesn't "
                f"produce a capturable result (TestCase ID={row_id_raw}). SaveAs is only valid on: "
                f"{sorted(SAVE_AS_CAPABLE)}"
            )

        data_source_col = col_idx.get("DataSource")  # optional column - see module docstring
        data_source = _clean(row[data_source_col]) if data_source_col is not None else ""

        step = TestStep(
            row_id=int(row_id_raw),
            test_scenario=_clean(row[col_idx["Test Scenario"]]),
            description=_clean(row[col_idx["Test Case Description"]]),
            keyword=keyword,
            locator_type=_clean(row[col_idx["LocatorType"]]),
            locator_value=locator_value,
            test_data=_clean(row[col_idx["Test Data"]]),
            expected_output=_clean(row[col_idx["Expected Output"]]),
            suite=_clean(row[col_idx["Suite"]]),
            save_as=save_as,
            data_source=data_source,
        )
        all_steps.append(step)

    all_steps.sort(key=lambda s: s.row_id)

    grouped = OrderedDict()
    for step in all_steps:
        grouped.setdefault(step.test_scenario, []).append(step)

    for scenario, steps in grouped.items():
        for position, step in enumerate(steps):
            if step.keyword.lower() == SESSION_DIRECTIVE_KEYWORD and position != 0:
                raise InvalidTestDataError(
                    f"Scenario '{scenario}': UseSession (Row {step.row_id}) must be the FIRST "
                    f"step of its Test Scenario - it configures the browser context before any "
                    f"other step runs, so it can't appear mid-case."
                )

    result = []
    for scenario, steps in grouped.items():
        if suite_filter:
            wanted = suite_filter.strip().lower()
            steps = [s for s in steps if wanted in _suite_list(s.suite)]
            if not steps:
                continue

        data_source = steps[0].data_source  # scenario-level attribute - only first row consulted
        if not data_source:
            result.append(TestCase(test_scenario=scenario, steps=steps))
            continue

        for row_dict in _load_data_table(wb, data_source):
            row_dict = dict(row_dict)  # don't mutate _load_data_table's own copy
            label = row_dict.pop("Label")
            result.append(TestCase(
                test_scenario=f"{scenario} [{label}]",
                steps=steps,
                data_row=row_dict,
            ))

    if suite_filter:
        logger.info(f"Filtered to Suite='{suite_filter}': {len(result)} test case(s)")
    logger.info(f"Loaded {len(result)} test case(s) from {path.name}")
    return result
